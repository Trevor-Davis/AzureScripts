#!/usr/bin/env python3
"""
Build an Excel pivot workbook from a CX Observe product/SKU hierarchy CSV.

Consumes the *_product_sku_hierarchy.csv produced by Get-CxoConsumption.ps1 -Hierarchy
and writes <same stem>_Pivot.xlsx containing:

  Pivot     collapsible Service -> Product -> SKU outline with % of parent / % of total
  Data      flat rows as a named Excel table (build your own PivotTable off this)
  Top SKUs  top N SKUs ranked, with parent product and service

Usage:
    python New-CxoPivot.py <hierarchy.csv> [-o output.xlsx] [--top 50]

The source data is Microsoft Confidential \\ Microsoft FTE. Label the workbook
before sharing it.
"""
import argparse
import csv
import os
import sys
from collections import OrderedDict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import DataBarRule
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
SVC_FILL = PatternFill("solid", fgColor="D9E2F3")
PRD_FILL = PatternFill("solid", fgColor="F2F2F2")
TOT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")

# Accounting format, whole units, no currency symbol - these are Azure Consumption
# Units (a normalized usage metric), not dollars. Keeps the aligned layout and the
# "-" for zero that Accounting gives you.
ACCT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'


def load(path):
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        rdr = csv.DictReader(fh)
        need = {"ServiceName_L2", "ProductName_L4", "SKU_L5", "SKU_ACU"}
        missing = need - set(rdr.fieldnames or [])
        if missing:
            sys.exit(f"{path} is missing expected column(s): {', '.join(sorted(missing))}")
        for r in rdr:
            raw = (r.get("SKU_ACU") or "").strip()
            try:
                acu = float(raw) if raw else 0.0
            except ValueError:
                acu = 0.0
            rows.append({
                "svc": (r["ServiceName_L2"] or "(unassigned)").strip(),
                "prod": (r["ProductName_L4"] or "").strip(),
                "sku": (r["SKU_L5"] or "").strip(),
                "acu": acu,
            })
    if not rows:
        sys.exit(f"No data rows found in {path}")
    return rows


def fmt_months(months):
    """'6 mo' / '1 mo' / '5.5 mo' - used in the ACU column headers."""
    if months is None:
        return None
    return f"{int(months)} mo" if float(months).is_integer() else f"{months:g} mo"


def build(rows, out_path, title, top_n, months):
    tree = OrderedDict()
    for r in rows:
        tree.setdefault(r["svc"], OrderedDict()).setdefault(r["prod"], []).append(r)

    svc_tot = {s: sum(x["acu"] for ps in p.values() for x in ps) for s, p in tree.items()}
    prod_tot = {(s, pr): sum(x["acu"] for x in sk)
                for s, p in tree.items() for pr, sk in p.items()}
    ordered_svcs = sorted(tree, key=lambda s: -svc_tot[s])

    label = fmt_months(months)
    acu_hdr = f"ACU ({label})" if label else "ACU"
    per_mo = bool(label)          # only offer a monthly column when we know the span

    wb = Workbook()

    # ---------------------------------------------------------------- Data
    # A Service | B Product | C SKU | D ACU | E ACU/mo | F % of Product | G % of Total
    ws = wb.active
    ws.title = "Data"
    hdr = ["Service (L2)", "Product (L4)", "SKU (L5)", acu_hdr]
    if per_mo:
        hdr.append("ACU / mo")
    hdr += ["% of Product", "% of Total"]
    ws.append(hdr)

    for s in ordered_svcs:
        for pr in sorted(tree[s], key=lambda p: -prod_tot[(s, p)]):
            for x in sorted(tree[s][pr], key=lambda v: -v["acu"]):
                ws.append([s, pr, x["sku"], x["acu"]] + ([None] if per_mo else []) + [None, None])

    n = ws.max_row
    c_mo = 5 if per_mo else None
    c_pct_prod = 6 if per_mo else 5
    c_pct_tot = 7 if per_mo else 6
    last_col = "G" if per_mo else "F"

    for i in range(2, n + 1):
        ws.cell(i, 4).number_format = ACCT
        if per_mo:
            ws.cell(i, c_mo).value = f"=D{i}/{months}"
            ws.cell(i, c_mo).number_format = ACCT
        ws.cell(i, c_pct_prod).value = f"=IFERROR(D{i}/SUMIFS($D$2:$D${n},$B$2:$B${n},$B{i}),0)"
        ws.cell(i, c_pct_tot).value = f"=IFERROR(D{i}/SUM($D$2:$D${n}),0)"
        ws.cell(i, c_pct_prod).number_format = "0.00%"
        ws.cell(i, c_pct_tot).number_format = "0.00%"

    tbl = Table(displayName="ConsumptionData", ref=f"A1:{last_col}{n}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)
    widths = [34, 42, 34, 20] + ([18] if per_mo else []) + [13, 13]
    for c, w in zip("ABCDEFG", widths):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"

    # ---------------------------------------------------------------- Pivot
    # A Name | B Level | C ACU | D ACU/mo | E % of Parent | F % of Total
    pv = wb.create_sheet("Pivot")
    pv["A1"] = title
    pv["A1"].font = Font(size=14, bold=True, color="1F3864")
    sub = "Azure Consumption Units (ACU). Source: CX Observe. Confidential \\ Microsoft FTE."
    if per_mo:
        sub = (f"Azure Consumption Units. ACU columns show the {label} total and the "
               f"monthly average (total / {months}). Source: CX Observe. "
               "Confidential \\ Microsoft FTE.")
    pv["A2"] = sub
    pv["A2"].font = Font(size=9, italic=True, color="595959")

    heads = ["Service / Product / SKU", "Level", acu_hdr]
    if per_mo:
        heads.append("ACU / mo")
    heads += ["% of Parent", "% of Total"]

    r0 = 4
    for j, h in enumerate(heads, 1):
        c = pv.cell(r0, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    ncols = len(heads)
    p_mo = 4 if per_mo else None
    p_parent = 5 if per_mo else 4
    p_total = 6 if per_mo else 5

    DL = f"Data!$D$2:$D${n}"
    SVC_RNG = f"Data!$A$2:$A${n}"
    PRD_RNG = f"Data!$B$2:$B${n}"

    r = r0 + 1
    total_row = r
    pv.cell(r, 1, "TOTAL").font = Font(bold=True, size=11)
    pv.cell(r, 2, "Total")
    pv.cell(r, 3, f"=SUM({DL})").font = Font(bold=True)
    if per_mo:
        pv.cell(r, p_mo, f"=C{r}/{months}").font = Font(bold=True)
    pv.cell(r, p_parent, 1)
    pv.cell(r, p_total, 1)
    for j in range(1, ncols + 1):
        pv.cell(r, j).fill = TOT_FILL
    r += 1

    for s in ordered_svcs:
        sr = r
        pv.cell(r, 1, s).font = Font(bold=True)
        pv.cell(r, 2, "Service")
        pv.cell(r, 3, f'=SUMIFS({DL},{SVC_RNG},$A{r})').font = Font(bold=True)
        if per_mo:
            pv.cell(r, p_mo, f"=C{r}/{months}")
        pv.cell(r, p_parent, f"=IFERROR(C{r}/$C${total_row},0)")
        pv.cell(r, p_total, f"=IFERROR(C{r}/$C${total_row},0)")
        for j in range(1, ncols + 1):
            pv.cell(r, j).fill = SVC_FILL
        r += 1

        for pr in sorted(tree[s], key=lambda p: -prod_tot[(s, p)]):
            prr = r
            safe_s = s.replace('"', '""')
            safe_p = pr.replace('"', '""')
            pv.cell(r, 1, "    " + pr).font = Font(bold=True, color="404040")
            pv.cell(r, 2, "Product")
            pv.cell(r, 3, f'=SUMIFS({DL},{SVC_RNG},"{safe_s}",{PRD_RNG},"{safe_p}")')
            if per_mo:
                pv.cell(r, p_mo, f"=C{r}/{months}")
            pv.cell(r, p_parent, f"=IFERROR(C{r}/C{sr},0)")
            pv.cell(r, p_total, f"=IFERROR(C{r}/$C${total_row},0)")
            for j in range(1, ncols + 1):
                pv.cell(r, j).fill = PRD_FILL
            pv.row_dimensions[r].outlineLevel = 1
            r += 1

            for x in sorted(tree[s][pr], key=lambda v: -v["acu"]):
                pv.cell(r, 1, "        " + x["sku"])
                pv.cell(r, 2, "SKU")
                pv.cell(r, 3, x["acu"])
                if per_mo:
                    pv.cell(r, p_mo, f"=C{r}/{months}")
                pv.cell(r, p_parent, f"=IFERROR(C{r}/C{prr},0)")
                pv.cell(r, p_total, f"=IFERROR(C{r}/$C${total_row},0)")
                pv.row_dimensions[r].outlineLevel = 2
                pv.row_dimensions[r].hidden = True
                r += 1

    last = r - 1
    for i in range(r0 + 1, r):
        pv.cell(i, 3).number_format = ACCT
        if per_mo:
            pv.cell(i, p_mo).number_format = ACCT
        pv.cell(i, p_parent).number_format = "0.0%"
        pv.cell(i, p_total).number_format = "0.0%"
        pv.cell(i, 2).alignment = Alignment(horizontal="center")
        for j in range(1, ncols + 1):
            pv.cell(i, j).border = Border(bottom=THIN)

    pv.sheet_properties.outlinePr.summaryBelow = False
    pv_widths = [56, 10, 20] + ([18] if per_mo else []) + [12, 12]
    for c, w in zip("ABCDEF", pv_widths):
        pv.column_dimensions[c].width = w
    pv.freeze_panes = "A5"
    pv.conditional_formatting.add(
        f"C{r0 + 2}:C{last}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="638EC6"))

    # ---------------------------------------------------------------- Top SKUs
    # A Rank | B SKU | C Product | D Service | E ACU | F ACU/mo | G % of Total
    ts = wb.create_sheet("Top SKUs")
    t_hdr = ["Rank", "SKU (L5)", "Product (L4)", "Service (L2)", acu_hdr]
    if per_mo:
        t_hdr.append("ACU / mo")
    t_hdr.append("% of Total")
    ts.append(t_hdr)
    for j in range(1, len(t_hdr) + 1):
        ts.cell(1, j).font = Font(bold=True, color="FFFFFF")
        ts.cell(1, j).fill = HDR_FILL

    t_mo = 6 if per_mo else None
    t_pct = 7 if per_mo else 6

    for i, x in enumerate(sorted(rows, key=lambda v: -v["acu"])[:top_n], 1):
        row = i + 1
        ts.append([i, x["sku"], x["prod"], x["svc"], x["acu"]]
                  + ([None] if per_mo else []) + [None])
        ts.cell(row, 5).number_format = ACCT
        if per_mo:
            ts.cell(row, t_mo).value = f"=E{row}/{months}"
            ts.cell(row, t_mo).number_format = ACCT
        ts.cell(row, t_pct).value = f"=IFERROR(E{row}/SUM(Data!$D$2:$D${n}),0)"
        ts.cell(row, t_pct).number_format = "0.00%"

    t_widths = [7, 36, 40, 30, 20] + ([18] if per_mo else []) + [12]
    for c, w in zip("ABCDEFG", t_widths):
        ts.column_dimensions[c].width = w
    ts.freeze_panes = "A2"
    if ts.max_row > 1:
        ts.conditional_formatting.add(
            f"E2:E{ts.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="max", color="63BE7B"))

    wb.save(out_path)
    return n - 1, last - r0, sum(r["acu"] for r in rows)


def infer_months(csv_path):
    """Derive the period length from the sibling *_all.csv StartDate/EndDate."""
    folder = os.path.dirname(csv_path)
    try:
        candidates = [f for f in os.listdir(folder) if f.endswith("_all.csv")]
    except OSError:
        return None
    for name in candidates:
        try:
            with open(os.path.join(folder, name), newline="", encoding="utf-8-sig") as fh:
                row = next(csv.DictReader(fh), None)
            if not row or "StartDate" not in row or "EndDate" not in row:
                continue
            start = datetime.strptime(row["StartDate"][:10], "%Y-%m-%d")
            end = datetime.strptime(row["EndDate"][:10], "%Y-%m-%d")
            months = (end.year - start.year) * 12 + (end.month - start.month)
            if months > 0:
                return float(months)
        except (OSError, ValueError, StopIteration):
            continue
    return None


def main():
    ap = argparse.ArgumentParser(description="Build a pivot workbook from a CX Observe hierarchy CSV.")
    ap.add_argument("csv_path", help="*_product_sku_hierarchy.csv")
    ap.add_argument("-o", "--output", help="output .xlsx (default: <stem>_Pivot.xlsx)")
    ap.add_argument("--top", type=int, default=50, help="rows on the Top SKUs sheet (default 50)")
    ap.add_argument("--title", help="title shown on the Pivot sheet")
    ap.add_argument("--months", type=float,
                    help="months the ACU totals cover; labels the ACU column and drives "
                         "the ACU / mo column. Inferred from the sibling *_all.csv if omitted.")
    args = ap.parse_args()

    src = os.path.abspath(args.csv_path)
    if not os.path.isfile(src):
        sys.exit(f"Not found: {src}")

    months = args.months if args.months else infer_months(src)
    if months is not None and months <= 0:
        months = None

    stem = os.path.basename(src)
    for suffix in ("_product_sku_hierarchy.csv", ".csv"):
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
            break

    out = args.output or os.path.join(os.path.dirname(src), f"{stem}_Pivot.xlsx")
    title = args.title or f"{stem.replace('_', ' ')} - Azure Consumption by Service > Product > SKU"

    data_rows, pivot_rows, grand = build(load(src), out, title, args.top, months)
    print(f"saved       : {out}")
    print(f"data rows   : {data_rows}")
    print(f"pivot rows  : {pivot_rows}")
    print(f"grand total : {grand:,.2f} ACU")
    if months:
        print(f"period      : {fmt_months(months)}  ({grand / months:,.2f} ACU/mo)")
    else:
        print("period      : unknown - ACU / mo column omitted (pass --months)")


if __name__ == "__main__":
    main()
